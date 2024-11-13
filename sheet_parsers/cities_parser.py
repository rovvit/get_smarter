import openpyxl
import random

COUNTRIES = [
    "Австрия",
    "Англия",
    "Беларусь",
    "Бельгия",
    "Гамельн",
    "Германия",
    "Греция",
    "Испания",
    "Италия",
    "Марокко",
    "Норвегия",
    "Польша",
    "Россия",
    "Сирия",
    "США",
    "Украина",
    "Франция",
    "Швеция"
]


def random_number(max):
    return random.randint(2, max)


def get_city_by_id(number):
    path = "C:\\Users\\User\\PycharmProjects\\get_smarter\\src\\Города.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    return [
        sheet_obj.cell(row=number, column=1).value,  # city
        sheet_obj.cell(row=number, column=2).value,  # country
        sheet_obj.cell(row=number, column=3).value  # comment
    ]


def generate_wrong_answers(correct_ans):
    answers = [correct_ans]
    while len(answers) < 4:
        rand = random_number(len(COUNTRIES))-1
        if COUNTRIES[rand] not in answers:
            answers.append(COUNTRIES[rand])
    return answers


def get_city_question():
    path = "C:\\Users\\User\\PycharmProjects\\get_smarter\\src\\Города.xlsx"
    wb_obj = openpyxl.load_workbook(path)
    sheet_obj = wb_obj.active
    number = random_number(sheet_obj.max_row)
    city = get_city_by_id(number)
    question = {
        "id": number,
        "question": city[0],
        "answer": city[1],
        "comment": city[2],
        "choices": generate_wrong_answers(city[1])
    }
    return question

# print(get_city_question())
